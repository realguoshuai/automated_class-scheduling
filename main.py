# Ensure local directory is in path for imports
import os
import sys
current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.insert(0, current_dir)

from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QTableWidget, QTableWidgetItem, 
                             QPushButton, QComboBox, QLabel, QHeaderView, 
                             QMessageBox, QFileDialog, QCheckBox)
from PyQt5.QtCore import Qt, QSize, QThread, pyqtSignal
from PyQt5.QtGui import QIcon, QFont, QPixmap
from datetime import datetime

from models import DataManager
from scheduler import Scheduler
from dialogs import DataManagementDialog, CourseSelectionDialog, HistoryDialog, TeacherStatisticsDialog

class SchedulingThread(QThread):
    finished = pyqtSignal(bool, str)
    
    def __init__(self, scheduler: Scheduler):
        super().__init__()
        self.scheduler = scheduler

    def run(self):
        try:
            success = self.scheduler.schedule()
            self.finished.emit(success, "")
        except Exception as e:
            self.finished.emit(False, str(e))

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.dm = DataManager()
        self.scheduler = Scheduler(self.dm)
        self.view_mode = "class" # "class" or "teacher"
        self.scheduling_thread = None
        
        self.setWindowTitle("自动化排课系统")
        self.resize(1400, 900)
        self.setMinimumSize(1024, 768)
        
        # Load Styles
        self.load_styles()
        
        # UI Components
        self.init_ui()
        
        # Initial Data Load
        self.update_entity_selector()
        self.refresh_timetable()

    def load_styles(self):
        if os.path.exists("styles.qss"):
            with open("styles.qss", "r", encoding="utf-8") as f:
                self.setStyleSheet(f.read())

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(15)
        
        # --- Sidebar ---
        sidebar = QWidget()
        sidebar.setFixedWidth(280)
        sidebar.setObjectName("sidebar")
        sidebar_layout = QVBoxLayout(sidebar)
        
        title_label = QLabel("控制面板")
        title_label.setObjectName("title_label")
        sidebar_layout.addWidget(title_label)
        
        # View Mode Selector
        sidebar_layout.addWidget(QLabel("展示维度:"))
        self.view_type_selector = QComboBox()
        self.view_type_selector.addItem("班级视角", "class")
        self.view_type_selector.addItem("教师视角", "teacher")
        self.view_type_selector.currentIndexChanged.connect(self.on_view_type_changed)
        sidebar_layout.addWidget(self.view_type_selector)

        # Entity Selector (Class or Teacher)
        sidebar_layout.addSpacing(10)
        self.entity_label = QLabel("当前班级:")
        sidebar_layout.addWidget(self.entity_label)
        self.entity_selector = QComboBox()
        self.entity_selector.currentIndexChanged.connect(self.refresh_timetable)
        sidebar_layout.addWidget(self.entity_selector)
        
        sidebar_layout.addSpacing(30)
        
        # Action Buttons
        self.manage_data_btn = QPushButton("数据管理")
        self.manage_data_btn.setObjectName("secondary_btn")
        self.manage_data_btn.clicked.connect(self.open_data_management)
        sidebar_layout.addWidget(self.manage_data_btn)
        
        self.run_scheduler_btn = QPushButton("一键自动排课")
        self.run_scheduler_btn.setObjectName("run_scheduler_btn")
        self.run_scheduler_btn.clicked.connect(self.run_auto_scheduling)
        sidebar_layout.addWidget(self.run_scheduler_btn)
        
        self.export_btn = QPushButton("导出 Excel")
        self.export_btn.setObjectName("secondary_btn")
        self.export_btn.clicked.connect(self.export_to_excel)
        sidebar_layout.addWidget(self.export_btn)

        self.screenshot_btn = QPushButton("保存课表为图片")
        self.screenshot_btn.setObjectName("secondary_btn")
        self.screenshot_btn.clicked.connect(self.take_screenshot)
        sidebar_layout.addWidget(self.screenshot_btn)
        
        self.history_btn = QPushButton("历史版本管理")
        self.history_btn.setObjectName("secondary_btn")
        self.history_btn.clicked.connect(self.open_history)
        sidebar_layout.addWidget(self.history_btn)
        
        self.stats_btn = QPushButton("🔍 查看排课统计")
        self.stats_btn.setObjectName("secondary_btn")
        self.stats_btn.clicked.connect(self.open_statistics)
        sidebar_layout.addWidget(self.stats_btn)
        
        self.export_global_btn = QPushButton("导出全校总表")
        self.export_global_btn.setObjectName("secondary_btn")
        self.export_global_btn.setStyleSheet("background-color: #28a745; color: white;") # Make it stand out
        self.export_global_btn.clicked.connect(self.export_global_excel)
        sidebar_layout.addWidget(self.export_global_btn)
        
        sidebar_layout.addSpacing(15)
        
        # Manual Mode Toggle
        self.manual_mode_checkbox = QCheckBox("开启手动调课 (双击单元格)")
        self.manual_mode_checkbox.setStyleSheet("color: #d73a49;") # Keep danger color
        sidebar_layout.addWidget(self.manual_mode_checkbox)

        sidebar_layout.addStretch()
        
        # --- Main Display Area ---
        display_area = QVBoxLayout()
        display_area.setSpacing(10)
        
        # Header Info
        self.info_label = QLabel("请选择对象查看课表")
        self.info_label.setObjectName("info_label")
        self.info_label.setStyleSheet("font-size: 24px; color: #1f2328; font-weight: bold; margin-bottom: 10px;")
        display_area.addWidget(self.info_label)
        
        # Timetable Grid
        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.cellDoubleClicked.connect(self.on_cell_double_clicked)
        display_area.addWidget(self.table)
        
        main_layout.addWidget(sidebar)
        main_layout.addLayout(display_area)
        
        # Status Bar
        self.statusBar().showMessage("系统就绪")

    def on_view_type_changed(self):
        self.view_mode = self.view_type_selector.currentData()
        self.entity_label.setText("当前班级:" if self.view_mode == "class" else "当前教师:")
        self.update_entity_selector()

    def update_entity_selector(self):
        self.entity_selector.blockSignals(True)
        self.entity_selector.clear()
        if self.view_mode == "class":
            for c in self.dm.classes:
                self.entity_selector.addItem(c.name, c.id)
        else:
            for t in self.dm.teachers:
                self.entity_selector.addItem(t.name, t.id)
        self.entity_selector.blockSignals(False)
        
        if self.entity_selector.count() > 0:
            self.refresh_timetable()

    def refresh_timetable(self):
        entity_id = self.entity_selector.currentData()
        if not entity_id:
            return
            
        if self.view_mode == "class":
            class_obj = self.dm.get_class_by_id(entity_id)
            if class_obj:
                self.info_label.setText(f"{class_obj.name} 课表")
            schedule = self.dm.timetable.get(entity_id, [])
        else:
            teacher_obj = self.dm.get_teacher_by_id(entity_id)
            if teacher_obj:
                self.info_label.setText(f"{teacher_obj.name} 老师课表")
            schedule = self.dm.get_teacher_timetable(entity_id)
        
        days = self.dm.settings["days_per_week"]
        periods = self.dm.total_periods
        
        self.table.setRowCount(periods)
        self.table.setColumnCount(days)
        
        # Headers
        day_names = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
        self.table.setHorizontalHeaderLabels(day_names[:days])
        
        # Generate row labels: Morning, Standard, Evening
        labels = []
        m_p = self.dm.morning_count
        s_p = self.dm.standard_count
        e_p = self.dm.evening_count
        
        for i in range(m_p): labels.append(f"早{i+1}")
        for i in range(s_p): labels.append(f"正{i+1}")
        for i in range(e_p): labels.append(f"晚{i+1}")
        self.table.setVerticalHeaderLabels(labels)
        
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        # Reset and Fill Data
        self.table.clearContents()
        if schedule:
            for r in range(min(len(schedule), periods)):
                for c in range(min(len(schedule[r]), days)):
                    item_data = schedule[r][c]
                    if isinstance(item_data, dict):
                        course_name = item_data.get("name", "")
                        teacher_name = item_data.get("teacher_name", "")
                        if self.view_mode == "class" and teacher_name:
                            display_text = f"{course_name} ({teacher_name})"
                        else:
                            display_text = course_name
                    elif isinstance(item_data, str):
                        display_text = item_data
                    
                    item = QTableWidgetItem(display_text)
                    item.setTextAlignment(Qt.AlignCenter)
                    
                    if self.view_mode == "class" and course_name:
                         # Visual grouping
                         if r < m_p: item.setBackground(Qt.cyan)
                         elif r >= m_p + s_p: item.setBackground(Qt.lightGray)
                         else: item.setBackground(Qt.white)
                    
                    self.table.setItem(r, c, item)

    def open_data_management(self):
        dialog = DataManagementDialog(self.dm, self)
        if dialog.exec_():
            self.dm.load_all() # Refresh data from disk
            self.update_entity_selector()
            self.refresh_timetable()
            self.statusBar().showMessage("数据已更新")

    def open_statistics(self):
        dialog = TeacherStatisticsDialog(self.dm, self)
        dialog.exec_()

    def open_history(self):
        dialog = HistoryDialog(self.dm, self)
        if dialog.exec_():
            self.update_entity_selector()
            self.refresh_timetable()
            self.statusBar().showMessage("历史版本已载入")

    def run_auto_scheduling(self):
        if self.scheduling_thread and self.scheduling_thread.isRunning():
            return

        # --- Pre-run check for gaps (Missing lessons) ---
        gaps = self.dm.check_curriculum_completeness()
        if gaps:
            gap_list = "\n".join([f" - {name}: 缺 {count} 节" for name, count in gaps[:10]])
            if len(gaps) > 10: gap_list += "\n..."
            reply = QMessageBox.warning(self, "课程设置不完整警示", 
                                       f"以下班级的课时总数不满 40 节，排完后会出现空时间段：\n\n{gap_list}\n\n提醒：建议去数据管理增加老师的授课课时。\n是否仍要继续排课？",
                                       QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.No:
                return

        reply = QMessageBox.question(self, "确认", "确定要重新排课吗？当前课表将被覆盖。",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            self.statusBar().showMessage("正在排课中，请稍候...")
            self.run_scheduler_btn.setEnabled(False)
            
            self.scheduling_thread = SchedulingThread(self.scheduler)
            self.scheduling_thread.finished.connect(self.handle_scheduling_finished)
            self.scheduling_thread.start()

    def handle_scheduling_finished(self, success, error_msg):
        self.run_scheduler_btn.setEnabled(True)
        if success:
            results = self.scheduler.get_result()
            self.dm.timetable = results
            # Auto save snapshot
            self.dm.save_snapshot("自动排课生成")
            self.dm.save_all()
            self.refresh_timetable()
            QMessageBox.information(self, "成功", "课表已自动生成！")
            self.statusBar().showMessage("排课成功")
        else:
            msg = error_msg if error_msg else "无法在当前约束下找到可行解。请尝试减少课时或放宽限制。"
            QMessageBox.warning(self, "失败", f"排课失败：{msg}")
            self.statusBar().showMessage("排课失败")

    def take_screenshot(self):
        # Create screenshots folder
        if not os.path.exists("screenshots"):
            os.makedirs("screenshots")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"screenshots/课表_{self.entity_selector.currentText()}_{timestamp}.png"
        
        pixmap = self.table.grab()
        pixmap.save(filename, "PNG")
        
        QMessageBox.information(self, "截图成功", f"课表截图已保存至：\n{os.path.abspath(filename)}")

    def on_cell_double_clicked(self, row, col):
        if not self.manual_mode_checkbox.isChecked():
            return
        
        if self.view_mode != "class":
            QMessageBox.warning(self, "提示", "请切换到'班级视角'后再进行调课。")
            return

        class_id = self.entity_selector.currentData()
        if not class_id: return

        # Get relevant courses for this class
        class_courses = [c for c in self.dm.courses if c.class_id == class_id]
        
        # Get current course in this cell
        class_schedule = self.dm.timetable.get(class_id, [])
        current_course_id = None
        if row < len(class_schedule) and col < len(class_schedule[row]):
            item = class_schedule[row][col]
            if isinstance(item, dict):
                current_course_id = item.get("course_id")

        dialog = CourseSelectionDialog(class_courses, current_course_id, self)
        if dialog.exec_():
            new_course_id = dialog.selected_course_id
            
            if new_course_id:
                # Validate Conflict
                can_place, msg = self.dm.check_slot_available(new_course_id, col, row)
                if not can_place:
                    QMessageBox.critical(self, "冲突警告", msg)
                    return
                
                # Update Timetable
                course = self.dm.get_course_by_id(new_course_id)
                new_data = {"course_id": new_course_id, "name": course.name}
            else:
                new_data = {"course_id": None, "name": ""}

            # Ensure grid is large enough
            days = self.dm.settings["days_per_week"]
            periods = self.dm.total_periods
            if class_id not in self.dm.timetable:
                self.dm.timetable[class_id] = [[{"course_id": None, "name": ""} for _ in range(days)] for _ in range(periods)]
            
            grid = self.dm.timetable[class_id]
            while len(grid) <= row: grid.append([{"course_id": None, "name": ""} for _ in range(days)])
            while len(grid[row]) <= col: grid[row].append({"course_id": None, "name": ""})
            
            grid[row][col] = new_data
            self.dm.save_all()
            self.refresh_timetable()
            self.statusBar().showMessage("手动调课成功")

    def export_to_excel(self):
        entity_id = self.entity_selector.currentData()
        if not entity_id:
            QMessageBox.warning(self, "提示", "请选择查看对象。")
            return
            
        path, _ = QFileDialog.getSaveFileName(self, "导出课表", f"{self.entity_selector.currentText()}课表.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return
            
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "课表"
            
            # Header
            days = self.dm.settings["days_per_week"]
            periods = self.dm.total_periods
            m_p = self.dm.morning_count
            s_p = self.dm.standard_count
            e_p = self.dm.evening_count
            
            day_names = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
            
            headers = ["节次"] + day_names[:days]
            ws.append(headers)
            
            # Formatting
            bold_font = Font(bold=True, color="FFFFFF")
            fill = PatternFill(start_color="24292E", end_color="24292E", fill_type="solid")
            alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in ws[1]:
                cell.font = bold_font
                cell.fill = fill
                cell.alignment = alignment
            
            # Data
            if self.view_mode == "class":
                schedule = self.dm.timetable.get(entity_id, [])
            else:
                schedule = self.dm.get_teacher_timetable(entity_id)

            # Row Labels
            labels = []
            for i in range(m_p): labels.append(f"早{i+1}")
            for i in range(s_p): labels.append(f"正{i+1}")
            for i in range(e_p): labels.append(f"晚{i+1}")

            for r in range(periods):
                row_data = [labels[r] if r < len(labels) else f"{r+1}"]
                for d in range(days):
                    course_str = ""
                    if r < len(schedule) and d < len(schedule[r]):
                        item = schedule[r][d]
                        if isinstance(item, dict):
                            course_name = item.get("name", "")
                            teacher_name = item.get("teacher_name", "")
                            if self.view_mode == "class" and teacher_name:
                                course_str = f"{course_name} ({teacher_name})"
                            else:
                                course_str = course_name
                        else:
                            course_str = str(item or "")
                    row_data.append(course_str)
                ws.append(row_data)
            
            # Style data cells
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
            
            for row in ws.iter_rows(min_row=2, max_row=periods+1, min_col=1, max_col=days+1):
                for cell in row:
                    cell.alignment = alignment
                    cell.border = border
            
            # Column width
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 15
                
            wb.save(path)
            QMessageBox.information(self, "成功", f"课表已成功导出至：\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败：{e}")

    def export_global_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "导出全校总表", "全校总课表.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return
            
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "全校总课表"
            
            # Setup dimensions
            days = self.dm.settings["days_per_week"]
            periods = self.dm.total_periods
            classes = self.dm.classes
            day_names = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
            
            # Row labels (Day + Period)
            labels = []
            m_p = self.dm.morning_count
            s_p = self.dm.standard_count
            e_p = self.dm.evening_count
            for i in range(m_p): labels.append(f"早{i+1}")
            for i in range(s_p): labels.append(f"正{i+1}")
            for i in range(e_p): labels.append(f"晚{i+1}")
            
            # Header Row
            headers = ["时间段"] + [c.name for c in classes]
            ws.append(headers)
            
            # Formatting
            header_fill = PatternFill(start_color="24292E", end_color="24292E", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center
                cell.border = border

            # Fill Data
            for d in range(days):
                day_name = day_names[d]
                for p in range(periods):
                    row_idx = d * periods + p
                    label = f"{day_name} - {labels[p]}"
                    row_data = [label]
                    
                    for cls in classes:
                        class_schedule = self.dm.timetable.get(cls.id, [])
                        course_str = ""
                        if p < len(class_schedule) and d < len(class_schedule[p]):
                            item = class_schedule[p][d]
                            if isinstance(item, dict) and item.get("name"):
                                course_id = item.get("course_id")
                                course = self.dm.get_course_by_id(course_id)
                                teacher_name = ""
                                if course:
                                    teacher = self.dm.get_teacher_by_id(course.teacher_id)
                                    teacher_name = f"({teacher.name})" if teacher else ""
                                course_str = f"{item['name']}{teacher_name}"
                        row_data.append(course_str)
                    
                    ws.append(row_data)
                    
                    # Style the row
                    excel_row = ws.max_row
                    # Color by day
                    day_fills = ["E8F5E9", "E3F2FD", "FFF3E0", "F3E5F5", "FBE9E7"]
                    fill = PatternFill(start_color=day_fills[d % len(day_fills)], 
                                      end_color=day_fills[d % len(day_fills)], fill_type="solid")
                    
                    for cell in ws[excel_row]:
                        cell.alignment = align_center
                        cell.border = border
                        if cell.column == 1:
                            cell.font = Font(bold=True)
                        else:
                            cell.fill = fill

            # Adjust Column Widths
            ws.column_dimensions['A'].width = 15
            for i in range(2, len(headers) + 1):
                col_letter = ws.cell(row=1, column=i).column_letter
                ws.column_dimensions[col_letter].width = 18
            
            # Freeze panes
            ws.freeze_panes = "B2"
            
            wb.save(path)
            QMessageBox.information(self, "成功", f"全校总表已导出至：\n{path}")
            
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"错误详情：{str(e)}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    # Set app-wide font for better Chinese display
    font = QFont("Microsoft YaHei", 9)
    app.setFont(font)
    
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
