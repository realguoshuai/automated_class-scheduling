import json
import os
import shutil
from dataclasses import dataclass, field, asdict
from typing import List, Dict, Tuple, Optional, Any

@dataclass
class SchoolClass:
    id: str
    name: str
    grade: str = ""

@dataclass
class Teacher:
    id: str
    name: str
    subject: str
    max_weekly: int
    target_grades: str = "" # e.g. "高一,高二"
    unavailable: List[str] = field(default_factory=list)  # Formatted as "周X第Y节"

@dataclass
class Course:
    id: str
    name: str
    teacher_id: str
    class_id: str
    weekly_hours: int
    consecutive: bool = False
    slot_type: str = "standard" # "morning", "standard", "evening"

class DataManager:
    def __init__(self, data_file="data.json", settings_file="settings.json"):
        self.data_file = data_file
        self.settings_file = settings_file
        self.classes: List[SchoolClass] = []
        self.teachers: List[Teacher] = []
        self.courses: List[Course] = []
        self.timetable: Dict[str, Any] = {}  # {class_id: [[course_id, ...], ...]}
        self.settings = {
            "school_name": "希望高级中学",
            "days_per_week": 5,
            "morning_periods": 1,
            "standard_periods": 8,
            "evening_periods": 3,
            "consecutive_allowed_starts": [1, 3, 5, 7]
        }
        self.load_all()

    def load_all(self):
        if os.path.exists(self.settings_file):
            with open(self.settings_file, 'r', encoding='utf-8') as f:
                self.settings.update(json.load(f))
        
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    # Normalize IDs during load to prevent mismatch
                    self.classes = [SchoolClass(id=str(c['id']).strip(), name=str(c['name']), grade=str(c.get('grade', ''))) for c in data.get("classes", [])]
                    self.teachers = [Teacher(id=str(t['id']).strip(), name=str(t['name']), subject=str(t['subject']), 
                                           max_weekly=int(t['max_weekly']), target_grades=str(t.get('target_grades', '')), 
                                           unavailable=t.get('unavailable', [])) for t in data.get("teachers", [])]
                    self.courses = [Course(id=str(co['id']).strip(), name=str(co['name']), teacher_id=str(co['teacher_id']).strip(), 
                                         class_id=str(co['class_id']).strip(), weekly_hours=int(co['weekly_hours']), 
                                         consecutive=co.get('consecutive', False), slot_type=str(co.get('slot_type', 'standard'))) 
                                         for co in data.get("courses", [])]
                    self.timetable = data.get("timetable", {})
            except Exception as e:
                print(f"Error loading data: {e}")

    def save_all(self):
        # Backup before save
        if os.path.exists(self.data_file):
            shutil.copy2(self.data_file, self.data_file + ".bak")

        data = {
            "classes": [asdict(c) for c in self.classes],
            "teachers": [asdict(t) for t in self.teachers],
            "courses": [asdict(co) for co in self.courses],
            "timetable": self.timetable
        }
        
        with open(self.data_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
            
        with open(self.settings_file, 'w', encoding='utf-8') as f:
            json.dump(self.settings, f, ensure_ascii=False, indent=2)

    def check_curriculum_completeness(self) -> List[Tuple[str, int]]:
        """Checks if all classes have exactly 40 standard hours. Returns list of (class_name, gap)"""
        target = self.standard_count * self.settings.get("days_per_week", 5)
        gaps = []
        for cls in self.classes:
            total = sum(co.weekly_hours for co in self.courses if co.class_id == cls.id and co.slot_type == "standard")
            if total < target:
                gaps.append((cls.name, target - total))
        return gaps

    @property
    def morning_count(self) -> int:
        val = self.settings.get("morning_periods", 0)
        return len(val) if isinstance(val, list) else int(val)

    @property
    def standard_count(self) -> int:
        val = self.settings.get("standard_periods", 8)
        return len(val) if isinstance(val, list) else int(val)

    @property
    def evening_count(self) -> int:
        val = self.settings.get("evening_periods", 0)
        return len(val) if isinstance(val, list) else int(val)

    @property
    def total_periods(self) -> int:
        return self.morning_count + self.standard_count + self.evening_count

    def get_teacher_by_id(self, teacher_id) -> Optional[Teacher]:
        for t in self.teachers:
            if t.id == teacher_id:
                return t
        return None

    def get_class_by_id(self, class_id) -> Optional[SchoolClass]:
        for c in self.classes:
            if c.id == class_id:
                return c
        return None

    def get_course_by_id(self, course_id) -> Optional[Course]:
        # Robust lookup with stripping
        target_id = str(course_id).strip()
        for co in self.courses:
            if str(co.id).strip() == target_id:
                return co
        return None

    def generate_unique_id(self, prefix: str) -> str:
        """Generates a unique ID based on existing items in the data manager."""
        existing_ids = set()
        if prefix == "T": existing_ids = {t.id for t in self.teachers}
        elif prefix == "C": existing_ids = {c.id for c in self.classes}
        elif prefix == "CO": existing_ids = {co.id for co in self.courses}
        
        i = 1
        while f"{prefix}_{i:03d}" in existing_ids:
            i += 1
        return f"{prefix}_{i:03d}"

    def check_slot_available(self, course_id, day, period) -> Tuple[bool, str]:
        course = self.get_course_by_id(course_id)
        if not course: return False, "课程不存在"

        # Check teacher unavailable
        teacher = self.get_teacher_by_id(course.teacher_id)
        if teacher:
            # Check grade restriction rules
            class_obj = self.get_class_by_id(course.class_id)
            if class_obj and class_obj.grade and teacher.target_grades:
                allowed_grades = [g.strip() for g in teacher.target_grades.split(",")]
                if class_obj.grade not in allowed_grades:
                    return False, f"配置规则拦截：教师 {teacher.name} 的授课年级（{teacher.target_grades}）不包含该班级所属的 {class_obj.grade}"
                    
            day_map = {0:"一", 1:"二", 2:"三", 3:"四", 4:"五", 5:"六", 6:"日"}
            unavail_str = f"周{day_map.get(day, '')}第{period+1}节"
            if unavail_str in teacher.unavailable:
                return False, f"{teacher.name}在此时段不可排课"

        # Check existing timetable for conflict
        for c_id, grid in self.timetable.items():
            if period < len(grid) and day < len(grid[period]):
                item = grid[period][day]
                existing_course_id = item.get("course_id") if isinstance(item, dict) else None
                if not existing_course_id:
                    continue
                
                existing_course = self.get_course_by_id(existing_course_id)
                if not existing_course:
                    continue
                
                # Check class conflict
                if c_id == course.class_id and existing_course_id != course_id:
                     return False, f"该班级在此时段已安排了{existing_course.name}"
                
                # Check teacher conflict (Teacher is teaching another class)
                if existing_course.teacher_id == course.teacher_id and existing_course_id != course_id:
                     conflict_class = self.get_class_by_id(c_id)
                     c_name = conflict_class.name if conflict_class else c_id
                     return False, f"{teacher.name if teacher else '该教师'}在此时段已在{c_name}上课"
        
        return True, ""

    def get_teacher_timetable(self, teacher_id) -> List[List[Dict]]:
        days = self.settings["days_per_week"]
        periods = self.total_periods
        
        teacher_grid = [[{"course_id": None, "name": ""} for _ in range(days)] for _ in range(periods)]
        
        for c_id, grid in self.timetable.items():
            class_obj = self.get_class_by_id(c_id)
            c_name = class_obj.name if class_obj else "未知班级"
            
            for p in range(min(periods, len(grid))):
                for d in range(min(days, len(grid[p]))):
                    item = grid[p][d]
                    if isinstance(item, dict) and item.get("course_id"):
                        course = self.get_course_by_id(item["course_id"])
                        if course and course.teacher_id == teacher_id:
                            teacher_grid[p][d] = {
                                "course_id": course.id,
                                "name": f"{course.name}\n({c_name})"
                            }
        return teacher_grid

    def generate_template(self, output_path):
        from openpyxl import Workbook
        wb = Workbook()
        
        # Teachers Sheet
        ws1 = wb.active
        ws1.title = "教师表"
        ws1.append(["ID", "姓名", "学科", "授课年级(逗号分隔,如 高一,高二)", "周最大课时", "不可排时间(周X第Y节，逗号分隔)"])
        ws1.append(["T001", "张老师", "语文", "高一", 14, "周一第1节,周二第2节"])
        
        # Classes Sheet
        ws2 = wb.create_sheet("班级表")
        ws2.append(["ID", "班级名称", "年级"])
        ws2.append(["C101", "高一(1)班", "高一"])
        
        # Courses Sheet
        ws3 = wb.create_sheet("科目设置表")
        ws3.append(["ID", "课程名", "教师ID", "班级ID", "周课时", "是否连堂(0/1)", "时段(morning/standard/evening)"])
        ws3.append(["CO101_CH", "语文", "T001", "C101", 6, 0, "standard"])
        ws3.append(["CO101_MOR", "早读", "T001", "C101", 3, 0, "morning"])
        
        wb.save(output_path)

    def import_from_excel(self, file_path):
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        
        # Import Teachers
        if "教师表" in wb.sheetnames:
            ws = wb["教师表"]
            self.teachers = []
            for row in list(ws.rows)[1:]:
                if row[0].value:
                    unavail = str(row[5].value).split(",") if (len(row) > 5 and row[5].value) else []
                    unavail = [u.strip() for u in unavail if u.strip()]
                    self.teachers.append(Teacher(
                        str(row[0].value), str(row[1].value), str(row[2].value),
                        int(row[4].value or 12), str(row[3].value or ""), unavail
                    ))

        # Import Classes
        if "班级表" in wb.sheetnames:
            ws = wb["班级表"]
            self.classes = []
            for row in list(ws.rows)[1:]:
                if row[0].value:
                    self.classes.append(SchoolClass(
                        str(row[0].value), str(row[1].value), str(row[2].value or "")
                    ))

        # Import Courses
        if "科目设置表" in wb.sheetnames:
            ws = wb["科目设置表"]
            self.courses = []
            for row in list(ws.rows)[1:]:
                if row[0].value:
                    self.courses.append(Course(
                        str(row[0].value), str(row[1].value or ""), str(row[2].value), 
                        str(row[3].value), int(row[4].value or 1), 
                        str(row[5].value) == "1", str(row[6].value or "standard")
                    ))
        
        self.save_all()

    def save_snapshot(self, name="自动存档"):
        import datetime
        history_dir = "history"
        if not os.path.exists(history_dir):
            os.makedirs(history_dir)
            
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"snapshot_{timestamp}.json"
        filepath = os.path.join(history_dir, filename)
        
        snapshot_data = {
            "timestamp": timestamp,
            "name": name,
            "timetable": self.timetable,
            "classes": [asdict(c) for c in self.classes],
            "teachers": [asdict(t) for t in self.teachers],
            "courses": [asdict(co) for co in self.courses]
        }
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(snapshot_data, f, ensure_ascii=False, indent=2)
        return filename

    def get_history_list(self) -> List[Dict]:
        history_dir = "history"
        if not os.path.exists(history_dir):
            return []
            
        history = []
        for filename in os.listdir(history_dir):
            if filename.startswith("snapshot_") and filename.endswith(".json"):
                try:
                    filepath = os.path.join(history_dir, filename)
                    with open(filepath, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        history.append({
                            "filename": filename,
                            "timestamp": data.get("timestamp", ""),
                            "name": data.get("name", "未知版本")
                        })
                except:
                    continue
        # Sort by timestamp descending
        history.sort(key=lambda x: x["timestamp"], reverse=True)
        return history

    def restore_snapshot(self, filename):
        history_dir = "history"
        filepath = os.path.join(history_dir, filename)
        if not os.path.exists(filepath):
            return False
            
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
            self.timetable = data.get("timetable", {})
            # We also restore classes/teachers/courses for full state consistency
            self.classes = [SchoolClass(**c) for c in data.get("classes", [])]
            self.teachers = [Teacher(**t) for t in data.get("teachers", [])]
            self.courses = [Course(**co) for co in data.get("courses", [])]
            self.save_all()
        return True
